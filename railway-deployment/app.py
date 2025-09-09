import os
from flask import Flask, render_template, request, jsonify, redirect, url_for
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from datetime import datetime
from werkzeug.utils import secure_filename
from openpyxl import load_workbook

from config import config
from database import db, DatabaseManager  
from auth import AuthManager, login_required, role_required

# Create Flask app
app = Flask(__name__)
config_name = os.getenv('FLASK_ENV', 'production')
app.config.from_object(config[config_name])

# Initialize rate limiter
limiter = Limiter(
    key_func=get_remote_address,
    default_limits=["100 per hour"]
)
limiter.init_app(app)

# Create directories
os.makedirs('data', exist_ok=True)
os.makedirs('uploads', exist_ok=True)

# Initialize database
try:
    db_manager = DatabaseManager()
    db_manager.init_database()
    print("Database initialized successfully")
except Exception as e:
    print(f"Database initialization error: {e}")

@app.route('/')
def index():
    from flask import session
    # Simple session check
    if session.get('logged_in'):
        return render_template('dashboard.html')
    return redirect(url_for('login'))



@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        data = request.get_json()
        email = data.get('email', '').strip().lower()
        password = data.get('password', '')
        
        # SIMPLE LOGIN - just use hardcoded credentials
        if email == 'admin@admin.com' and password == 'admin123':
            # Set simple session
            from flask import session
            session['user_id'] = 1
            session['user_role'] = 'admin'
            session['user_email'] = email
            session['user_name'] = 'Administrator'
            session['logged_in'] = True
            
            return jsonify({
                'success': True,
                'message': 'Login successful',
                'user': {
                    'id': 1,
                    'email': email,
                    'role': 'admin',
                    'name': 'Administrator'
                }
            })
        else:
            return jsonify({'success': False, 'message': 'Invalid credentials'}), 401
    
    return render_template('login.html')

@app.route('/api/employees', methods=['GET'])
@login_required
def get_employees():
    try:
        with db.get_connection() as conn:
            # Count total employees first
            count_cursor = conn.execute('SELECT COUNT(*) as total FROM employees WHERE is_active = 1')
            total = count_cursor.fetchone()['total']
            print(f"DEBUG: Found {total} active employees")
            
            cursor = conn.execute('''
                SELECT id, name, email, phone, department, position, 
                       hire_date, photo_path, total_entries, is_active,
                       created_at, updated_at
                FROM employees WHERE is_active = 1
                ORDER BY name
            ''')
            
            employees = []
            for row in cursor.fetchall():
                employee = dict(row)
                employees.append(employee)
                print(f"DEBUG: Added employee: {employee['name']}")
            
            response_data = {
                'success': True,
                'employees': employees,
                'total': len(employees)
            }
            print(f"DEBUG: Returning {len(employees)} employees")
            
            return jsonify(response_data)
            
    except Exception as e:
        print(f"DEBUG: Error in get_employees: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/employee', methods=['POST'])
@login_required
@role_required('manager')
def add_employee():
    try:
        data = request.get_json()
        name = data.get('name', '').strip()
        print(f"DEBUG: Adding employee: '{name}'")
        
        if not name:
            return jsonify({'success': False, 'error': 'Name is required'}), 400
        
        with db.get_connection() as conn:
            cursor = conn.execute(
                'SELECT id FROM employees WHERE name = ? AND is_active = 1', 
                (name,)
            )
            if cursor.fetchone():
                print(f"DEBUG: Employee '{name}' already exists")
                return jsonify({'success': False, 'error': 'Employee already exists'}), 400
            
            conn.execute('''
                INSERT INTO employees (name, total_entries, is_active, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?)
            ''', (name, 0, 1, datetime.now(), datetime.now()))
            
            conn.commit()
            print(f"DEBUG: Successfully added employee '{name}'")
            
            return jsonify({
                'success': True,
                'message': f'Employee "{name}" added successfully'
            })
            
    except Exception as e:
        print(f"DEBUG: Error adding employee: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/employee/<int:employee_id>/add_entry', methods=['POST'])
@login_required
@role_required('manager')
def add_entry(employee_id):
    try:
        data = request.get_json()
        activity_name = data.get('activity_name', '').strip()
        entries_awarded = int(data.get('entries_awarded', 1))
        
        if not activity_name:
            return jsonify({'success': False, 'error': 'Activity name is required'}), 400
        
        if entries_awarded <= 0 or entries_awarded > 10:
            return jsonify({'success': False, 'error': 'Entries must be between 1 and 10'}), 400
        
        with db.get_connection() as conn:
            # Check if employee exists
            cursor = conn.execute('SELECT id, name, total_entries FROM employees WHERE id = ? AND is_active = 1', (employee_id,))
            employee = cursor.fetchone()
            
            if not employee:
                return jsonify({'success': False, 'error': 'Employee not found'}), 404
            
            # Add activity
            conn.execute('''
                INSERT INTO activities (employee_id, activity_name, activity_category, 
                                      entries_awarded, awarded_by)
                VALUES (?, ?, ?, ?, ?)
            ''', (employee_id, activity_name, 'manual', entries_awarded, request.current_user['id']))
            
            # Update employee total entries
            new_total = employee['total_entries'] + entries_awarded
            conn.execute('UPDATE employees SET total_entries = ?, updated_at = ? WHERE id = ?', 
                        (new_total, datetime.now(), employee_id))
            
            conn.commit()
            
            return jsonify({
                'success': True,
                'message': f'Added {entries_awarded} entries for {employee["name"]}'
            })
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/employee/<int:employee_id>/reset_points', methods=['POST'])
@login_required
@role_required('manager')
def reset_points(employee_id):
    try:
        with db.get_connection() as conn:
            cursor = conn.execute('SELECT name FROM employees WHERE id = ? AND is_active = 1', (employee_id,))
            employee = cursor.fetchone()
            
            if not employee:
                return jsonify({'success': False, 'error': 'Employee not found'}), 404
            
            # Reset points
            conn.execute('UPDATE employees SET total_entries = 0, updated_at = ? WHERE id = ?', 
                        (datetime.now(), employee_id))
            
            # Remove all activities for this employee
            conn.execute('DELETE FROM activities WHERE employee_id = ?', (employee_id,))
            
            conn.commit()
            
            return jsonify({
                'success': True,
                'message': f'Reset points for {employee["name"]}'
            })
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/import_excel', methods=['POST'])
@login_required
@role_required('manager')
def import_excel():
    try:
        print("DEBUG: Excel import started")
        
        if 'file' not in request.files:
            print("DEBUG: No file in request")
            return jsonify({'success': False, 'error': 'No file provided'}), 400
        
        file = request.files['file']
        print(f"DEBUG: File received: {file.filename}")
        
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'error': 'Invalid file type. Please upload Excel files only.'}), 400
        
        # Process Excel file
        print("DEBUG: Loading workbook...")
        workbook = load_workbook(file)
        sheet = workbook.active
        print(f"DEBUG: Sheet loaded, max row: {sheet.max_row}")
        
        employees_added = 0
        errors = []
        names_found = []
        
        with db.get_connection() as conn:
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not any(row):
                    continue
                
                # Try to find and combine first and last names
                name = None
                first_name = None
                last_name = None
                
                # Check for separate first and last name columns
                for col_idx, cell in enumerate(row[:10]):  # Check first 10 columns
                    if cell and isinstance(cell, str) and len(cell.strip()) > 1:
                        cell_value = cell.strip()
                        
                        # Look for first name column
                        if col_idx == 0 or 'first' in str(sheet.cell(1, col_idx + 1).value or '').lower():
                            first_name = cell_value
                            print(f"DEBUG: Found first name '{first_name}' in column {col_idx}")
                        
                        # Look for last name column
                        elif col_idx == 1 or 'last' in str(sheet.cell(1, col_idx + 1).value or '').lower():
                            last_name = cell_value
                            print(f"DEBUG: Found last name '{last_name}' in column {col_idx}")
                        
                        # If it looks like a full name (has space)
                        elif ' ' in cell_value and len(cell_value.split()) >= 2:
                            name = cell_value
                            print(f"DEBUG: Found full name '{name}' in column {col_idx}")
                            break
                        
                        # Single name fallback
                        elif not first_name and len(cell_value) > 2:
                            first_name = cell_value
                            print(f"DEBUG: Using '{first_name}' as name from column {col_idx}")
                
                # Combine first and last names if found separately
                if first_name and last_name:
                    name = f"{first_name} {last_name}"
                    print(f"DEBUG: Combined name: '{name}'")
                elif first_name:
                    name = first_name
                    print(f"DEBUG: Using first name only: '{name}'")
                
                if not name:
                    continue
                    
                names_found.append(name)
                
                try:
                    # Check if employee exists
                    cursor = conn.execute('SELECT id FROM employees WHERE name = ? AND is_active = 1', (name,))
                    if not cursor.fetchone():
                        # Add employee
                        conn.execute('''
                            INSERT INTO employees (name, total_entries, is_active, created_at, updated_at)
                            VALUES (?, ?, ?, ?, ?)
                        ''', (name, 0, 1, datetime.now(), datetime.now()))
                        employees_added += 1
                        print(f"DEBUG: Added employee: {name}")
                    else:
                        print(f"DEBUG: Employee already exists: {name}")
                        
                except Exception as e:
                    error_msg = f"Row {row_num}: {str(e)}"
                    errors.append(error_msg)
                    print(f"DEBUG: Error adding employee: {error_msg}")
            
            conn.commit()
            print(f"DEBUG: Committed {employees_added} new employees")
        
        result = {
            'success': True,
            'message': f'Import completed. Added {employees_added} employees.',
            'employees_added': employees_added,
            'names_found': names_found,
            'errors': errors
        }
        print(f"DEBUG: Import result: {result}")
        
        return jsonify(result)
        
    except Exception as e:
        error_msg = f'Import failed: {str(e)}'
        print(f"DEBUG: Import exception: {error_msg}")
        return jsonify({'success': False, 'error': error_msg}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)